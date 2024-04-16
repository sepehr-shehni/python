from openpyxl import Workbook, load_workbook

# ایجاد یک شیء Workbook جدید
wb = Workbook()

# ایجاد یک شیت جدید
ws1 = wb.active
ws1.title = "Sheet1"

# افزودن داده به شیت
ws1['A1'] = 'Name'
ws1['B1'] = 'Age'
ws1['A2'] = 'John'
ws1['B2'] = 30
ws1['A3'] = 'Alice'
ws1['B3'] = 25

# ایجاد یک شیت دیگر
ws2 = wb.create_sheet(title="Sheet2")

# افزودن داده به شیت دوم
for row in range(1, 6):
    ws2.cell(row=row, column=1, value=f"Data {row}")

# ذخیره فایل Excel
wb.save(filename='example.xlsx')

# بارگذاری فایل Excel موجود
wb_loaded = load_workbook('example.xlsx')

# دسترسی به شیت فعلی
ws_loaded = wb_loaded.active

# ایجاد یک شیت جدید
ws3 = wb_loaded.create_sheet(title="Sheet3")

# بازگرداندن نام تمام شیت‌ها
sheet_names = wb_loaded.sheetnames
print(sheet_names)
